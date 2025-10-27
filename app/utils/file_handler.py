import os
import aiofiles
import time
import PyPDF2
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class FileHandler:
    """Handle file operations for the OCRD extractor"""

    def __init__(self):
        # ✅ Use /tmp for Render (ephemeral but writable)
        self.upload_dir = "/tmp/uploads"
        self.export_dir = "/tmp/exports"
        self._ensure_directories()

    def _ensure_directories(self):
        """Ensure required directories exist"""
        os.makedirs(self.upload_dir, exist_ok=True)
        os.makedirs(self.export_dir, exist_ok=True)

    async def save_uploaded_file(self, file) -> str:
        """Save uploaded file and return file path"""
        # Generate unique filename
        safe_name = file.filename.replace("/", "_").replace("\\", "_")
        filename = f"{int(time.time())}_{safe_name}"
        file_path = os.path.join(self.upload_dir, filename)

        # Save file asynchronously
        async with aiofiles.open(file_path, "wb") as f:
            content = await file.read()
            await f.write(content)

        return file_path

    async def extract_pdf_text(self, file_path: str) -> str:
        """Extract text from PDF file"""
        try:
            with open(file_path, "rb") as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page_num in range(len(pdf_reader.pages)):
                    page = pdf_reader.pages[page_num]
                    text += page.extract_text() or ""
                    text += "\n"
                return text.strip()
        except Exception as e:
            raise Exception(f"Failed to extract text from PDF: {str(e)}")

    async def create_excel_export(self, data: dict) -> str:
        """Create Excel export from analysis results"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "OCRD Results"

            # Define styles
            header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
            white_font = Font(color="FFFFFF", bold=True)
            center_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Add headers
            headers = ["Section", "Instrument", "Allowed", "Note", "Evidence"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = white_font
                cell.alignment = center_alignment
                cell.border = thin_border

            # Add data
            row = 2
            for section, items in data.get("sections", {}).items():
                for key, value in items.items():
                    if isinstance(value, dict) and "allowed" in value:
                        ws.cell(row=row, column=1, value=section)
                        ws.cell(row=row, column=2, value=key)
                        ws.cell(row=row, column=3, value="✓" if value.get("allowed") else "")
                        ws.cell(row=row, column=4, value=value.get("note", ""))
                        # Only show evidence for allowed items
                        evidence_text = ""
                        if value.get("allowed"):
                            evidence_text = value.get("evidence", {}).get("text", "")
                        ws.cell(row=row, column=5, value=evidence_text)

                        for col in range(1, 6):
                            ws.cell(row=row, column=col).border = thin_border
                        row += 1

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            # ✅ Save under /tmp
            filename = f"ocrd_export_{int(time.time())}.xlsx"
            file_path = os.path.join(self.export_dir, filename)
            wb.save(file_path)
            return file_path

        except Exception as e:
            raise Exception(f"Failed to create Excel export: {str(e)}")

    def cleanup_file(self, file_path: str):
        """Clean up temporary file"""
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
        except Exception:
            pass
