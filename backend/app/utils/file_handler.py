import os
import tempfile
import aiofiles
import time
import subprocess
import json
import re
import unicodedata
from typing import Optional, List, Dict, Any
from pathlib import Path
from uuid import uuid4
import PyPDF2
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .trace_handler import TraceHandler
from ..services.rag_index import index_pdf
from .logger import setup_logger

logger = setup_logger(__name__)

# Try to import NLTK for sentence tokenization
try:
    import nltk
    from nltk.tokenize import sent_tokenize
    NLTK_AVAILABLE = True
    # Download required NLTK data
    try:
        nltk.data.find('tokenizers/punkt')
    except LookupError:
        nltk.download('punkt', quiet=True)
except ImportError:
    NLTK_AVAILABLE = False

# Try to import LangChain for improved chunking
try:
    from langchain_text_splitters import RecursiveCharacterTextSplitter
    LANGCHAIN_AVAILABLE = True
except ImportError:
    LANGCHAIN_AVAILABLE = False

# Try to import optional dependencies
try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False

try:
    from pdfminer.high_level import extract_text as pdfminer_extract
    PDFMINER_AVAILABLE = True
except ImportError:
    PDFMINER_AVAILABLE = False

try:
    import camelot
    CAMELOT_AVAILABLE = True
except ImportError:
    CAMELOT_AVAILABLE = False

try:
    from pdf2image import convert_from_path
    PDF2IMAGE_AVAILABLE = True
except ImportError:
    PDF2IMAGE_AVAILABLE = False

try:
    from pypdf import PdfReader
    PYPDF_AVAILABLE = True
except ImportError:
    PYPDF_AVAILABLE = False

class FileHandler:
    """Handle file operations for the OCRD extractor"""
    
    # Negation cues pattern for preserving exceptions and negations
    NEG_CUES = re.compile(r"\b(not|no|except|unless|excluded|exclusion|prohibit|forbidden|restricted|ban(ned)?|provided\s+that|excluding|not\s+permitted|not\s+allowed)\b", re.IGNORECASE)
    
    def __init__(self):
        self.upload_dir = "uploads"
        self.export_dir = "exports"
        self.markdown_dir = "markdown"
        self.trace_handler = TraceHandler()
        self._ensure_directories()
    
    def _ensure_directories(self):
        """Ensure required directories exist"""
        os.makedirs(self.upload_dir, exist_ok=True)
        os.makedirs(self.export_dir, exist_ok=True)
        os.makedirs(self.markdown_dir, exist_ok=True)
    
    def is_image_only_pdf(self, file_path: str) -> bool:
        """
        Detect if PDF is image-only (scanned) by checking if extractable text is minimal.
        
        This method is conservative - it only flags as image-only if:
        - Total text is very low (< 100 chars per page on average)
        - AND at least 3 pages have been checked
        - OR total text is extremely low (< 200 chars total)
        
        Args:
            file_path: Path to PDF file
            
        Returns:
            True if PDF appears to be image-only (scanned), False otherwise
        """
        try:
            if not PYPDF_AVAILABLE:
                # Fallback: try PyPDF2
                with open(file_path, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    total_pages = len(pdf_reader.pages)
                    # Check first few pages and last page to get a better sample
                    pages_to_check = min(5, total_pages)
                    text_parts = []
                    for i in range(pages_to_check):
                        text_parts.append(pdf_reader.pages[i].extract_text() or "")
                    # Also check last page if we have more than 5 pages
                    if total_pages > 5:
                        text_parts.append(pdf_reader.pages[-1].extract_text() or "")
                    text = "".join(text_parts)
            else:
                reader = PdfReader(file_path)
                total_pages = len(reader.pages)
                # Check first few pages and last page to get a better sample
                pages_to_check = min(5, total_pages)
                text_parts = []
                for i in range(pages_to_check):
                    text_parts.append(reader.pages[i].extract_text() or "")
                # Also check last page if we have more than 5 pages
                if total_pages > 5:
                    text_parts.append(reader.pages[-1].extract_text() or "")
                text = "".join(text_parts)
            
            text_length = len(text.strip())
            
            # Very conservative detection:
            # - If total text is extremely low (< 200 chars), likely image-only
            # - OR if we checked multiple pages and average is very low (< 100 chars per page)
            if text_length < 200:
                return True
            
            # If we checked multiple pages, check average
            pages_checked = min(5, total_pages) + (1 if total_pages > 5 else 0)
            if pages_checked >= 3:
                avg_chars_per_page = text_length / pages_checked
                if avg_chars_per_page < 100:
                    return True
            
            # Default: assume it's a text PDF
            return False
        except Exception as e:
            logger.warning(f"Error checking if PDF is image-only: {e}")
            # If we can't check, assume it's not image-only and let normal pipeline handle it
            return False
    
    async def save_uploaded_file(self, file, max_bytes: int = 100 * 1024 * 1024) -> str:
        """
        Save UploadFile to disk without loading it all into memory.
        max_bytes = hard limit to protect RAM/disk (default 100MB).
        """
        filename = f"{int(time.time())}_{file.filename}"
        file_path = os.path.join(self.upload_dir, filename)

        total = 0
        chunk_size = 1024 * 1024  # 1MB

        async with aiofiles.open(file_path, "wb") as out:
            while True:
                chunk = await file.read(chunk_size)
                if not chunk:
                    break
                total += len(chunk)
                if total > max_bytes:
                    # cleanup partial file
                    try:
                        await out.close()
                    finally:
                        if os.path.exists(file_path):
                            os.remove(file_path)
                    raise ValueError(f"File too large (>{max_bytes} bytes)")
                await out.write(chunk)

        await file.close()
        return file_path
    
    async def extract_pdf_text(self, file_path: str, max_pages: Optional[int] = None) -> str:
        """Extract text from PDF file - extracts ALL pages"""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                total_pages = len(pdf_reader.pages)
                
                logger.info(f"Extracting text from all {total_pages} pages...")
                text = ""
                
                # Extract all pages, but process in batches to avoid memory issues
                for page_num in range(total_pages):
                    page = pdf_reader.pages[page_num]
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
                    
                    # Log progress for large documents
                    if (page_num + 1) % 25 == 0:
                        logger.info(f"Extracted {page_num + 1}/{total_pages} pages ({len(text)} chars so far)")
                
                logger.info(f"Extraction complete: {total_pages} pages, {len(text)} characters")
                return text.strip()
        except Exception as e:
            raise Exception(f"Failed to extract text from PDF: {str(e)}")
    
    async def extract_pdf_text_with_tracing(self, file_path: str, trace_id: str) -> Dict[str, Any]:
        """Extract text from PDF with robust fallback chain and forensic tracing"""
        try:
            start_time = time.time()
            trace_dir = Path(self.trace_handler.get_trace_dir(trace_id))
            
            # Initialize metadata
            meta = {
                "pdf_path": str(file_path),
                "trace_id": trace_id,
                "extraction_methods": [],
                "total_pages": 0,
                "char_count": 0,
                "tables_found": 0,
                "ocr_used": False,
                "is_image_only": False
            }
            
            # Always try text extraction first using robust fallback chain
            # This ensures normal text PDFs work correctly
            logger.info(f"📄 Attempting text extraction for: {file_path}")
            page_texts, extraction_methods = await self._extract_text_robust(file_path, trace_dir)
            
            # Validate extraction
            total_chars = sum(len(t) for t in page_texts)
            if total_chars == 0:
                logger.error(f"⚠️ CRITICAL: Text extraction returned 0 characters! File: {file_path}")
                logger.error(f"⚠️ Extraction methods tried: {extraction_methods}")
                # Try OCR as last resort if not already tried
                if not any(method.get("method") == "ocr" for method in extraction_methods):
                    logger.warning("⚠️ Attempting OCR as last resort...")
                    try:
                        ocr_texts = await self._extract_with_ocr(file_path, trace_dir)
                        if ocr_texts and sum(len(t) for t in ocr_texts) > 0:
                            page_texts = ocr_texts
                            extraction_methods.append({"method": "ocr", "char_count": sum(len(t) for t in ocr_texts), "success": True})
                            total_chars = sum(len(t) for t in page_texts)
                            logger.info(f"✅ OCR succeeded! Extracted {total_chars} characters")
                    except Exception as ocr_error:
                        logger.error(f"❌ OCR also failed: {ocr_error}")
                        extraction_methods.append({"method": "ocr", "error": str(ocr_error), "success": False})
            
            # Determine if PDF is truly image-only based on extraction results
            # Only mark as image-only if we got very little text after trying all methods
            # This is more reliable than pre-checking
            is_image_only = False
            if total_chars == 0:
                # No text extracted at all - likely image-only
                is_image_only = True
                logger.warning(f"📸 No text extracted - PDF appears to be image-only (scanned). Will need vision pipeline.")
            elif total_chars < 500 and len(page_texts) > 0:
                # Very little text relative to number of pages - likely image-only
                avg_chars_per_page = total_chars / len(page_texts)
                if avg_chars_per_page < 100:
                    is_image_only = True
                    logger.warning(f"📸 Very little text extracted ({total_chars} chars, {avg_chars_per_page:.1f} chars/page) - PDF appears to be image-only (scanned). Will need vision pipeline.")
            
            # Update metadata
            meta.update({
                "extraction_methods": extraction_methods,
                "total_pages": len(page_texts),
                "char_count": total_chars,
                "ocr_used": any(method.get("method") == "ocr" for method in extraction_methods),
                "is_image_only": is_image_only
            })
            
            # Save raw text for each page
            for i, page_text in enumerate(page_texts):
                await self.trace_handler.save_raw_text_page(
                    trace_id, i + 1, page_text
                )
            
            # Clean and normalize text
            clean_text = self._clean_text_robust(page_texts)
            
            # Warn if clean text is empty or very short
            if len(clean_text) == 0:
                logger.error(f"⚠️ CRITICAL: Clean text is empty after processing! Raw text had {total_chars} characters.")
                # If clean text is empty but we have pages, mark as image-only
                if len(page_texts) > 0:
                    is_image_only = True
            elif len(clean_text) < 100:
                logger.warning(f"⚠️ WARNING: Clean text is very short ({len(clean_text)} chars). PDF might be image-based or have extraction issues.")
            
            # Save clean text
            await self.trace_handler.save_clean_text(trace_id, clean_text)
            
            # Extract tables if available
            tables = await self._extract_tables(file_path, trace_dir)
            meta["tables_found"] = len(tables)
            
            # NEW: If tables/images are detected, prefer OCR-extracted text to preserve X/- marks
            has_tables = len(tables) > 0
            should_use_ocr = has_tables or is_image_only or len(clean_text) < 1000
            
            if should_use_ocr and not meta.get("ocr_used", False):
                logger.info(f"📸 Tables/images detected or low text quality - attempting OCR extraction to preserve X/- marks...")
                ocr_texts = await self._extract_with_ocr(file_path, trace_dir)
                if ocr_texts and sum(len(t) for t in ocr_texts) > len(clean_text) * 0.5:  # OCR gives at least 50% more text
                    ocr_clean_text = self._clean_text_robust(ocr_texts)
                    logger.info(f"✅ OCR extraction successful: {len(ocr_clean_text)} chars (vs {len(clean_text)} from regular extraction)")
                    clean_text = ocr_clean_text
                    meta["ocr_used"] = True
                    extraction_methods.append({"method": "ocr_preferred", "char_count": len(ocr_clean_text), "success": True, "reason": "tables/images detected"})
                else:
                    logger.info(f"⚠️ OCR extraction didn't improve text quality, using regular extraction")
            
            if tables:
                # Stitch tables back into text with clear markers
                clean_text = self._stitch_tables_into_text(clean_text, tables)
                # Save tables separately
                await self.trace_handler.save_tables(trace_id, tables)
            
            # Create chunks with improved LangChain-based chunking
            chunks = self.chunk_text(clean_text)
            await self.trace_handler.save_chunks(trace_id, chunks)
            
            # Save chunks in JSONL format for verification
            await self.save_chunks_jsonl(chunks, trace_id)
            
            # Index chunks for RAG retrieval
            trace_dir = self.trace_handler.get_trace_dir(trace_id)
            clean_text_path = os.path.join(trace_dir, "20_clean_text.txt")
            chunks_path = os.path.join(trace_dir, "30_chunks.jsonl")
            vectordb_dir = "var/chroma"
            
            # Perform RAG indexing
            rag_results = index_pdf(
                clean_text_path=clean_text_path,
                chunks_path=chunks_path,
                vectordb_dir=vectordb_dir,
                doc_id=trace_id
            )
            
            # Save RAG indexing results
            await self.trace_handler.save_rag_index(trace_id, rag_results)
            
            # Update final metadata with is_image_only status
            meta.update({
                "clean_text_length": len(clean_text),
                "chunks_count": len(chunks),
                "extraction_time": time.time() - start_time,
                "rag_indexed": rag_results.get("success", False),
                "rag_chunks_indexed": rag_results.get("indexed", 0),
                "is_image_only": is_image_only
            })
            
            # Save metadata
            await self.trace_handler.save_meta(trace_id, meta)
            
            extraction_time = time.time() - start_time
            
            return {
                "raw_text": "\n".join(page_texts),
                "clean_text": clean_text,
                "page_texts": page_texts,
                "chunks": chunks,
                "total_pages": len(page_texts),
                "extraction_time": extraction_time,
                "extraction_methods": extraction_methods,
                "ocr_used": meta["ocr_used"],
                "is_image_only": is_image_only
            }
                
        except Exception as e:
            raise Exception(f"Failed to extract text from PDF: {str(e)}")
    
    async def _extract_text_robust(self, file_path: str, trace_dir: Path) -> tuple[List[str], List[Dict]]:
        """Robust text extraction with fallback chain"""
        methods_used = []
        page_texts = []
        
        # Method 1: Try PyMuPDF (best for most PDFs)
        if PYMUPDF_AVAILABLE:
            try:
                page_texts, char_count = self._extract_with_pymupdf(file_path)
                methods_used.append({"method": "pymupdf", "char_count": char_count, "success": True})
                
                # Heuristic: if too little text, try OCR
                if char_count < 2000:
                    logger.info(f"Low text count ({char_count}), attempting OCR...")
                    ocr_texts = await self._extract_with_ocr(file_path, trace_dir)
                    if ocr_texts and sum(len(t) for t in ocr_texts) > char_count:
                        page_texts = ocr_texts
                        methods_used.append({"method": "ocr", "char_count": sum(len(t) for t in ocr_texts), "success": True})
                
                return page_texts, methods_used
            except Exception as e:
                methods_used.append({"method": "pymupdf", "error": str(e), "success": False})
        
        # Method 2: Try pdfminer
        if PDFMINER_AVAILABLE:
            try:
                page_texts, char_count = self._extract_with_pdfminer(file_path)
                methods_used.append({"method": "pdfminer", "char_count": char_count, "success": True})
                return page_texts, methods_used
            except Exception as e:
                methods_used.append({"method": "pdfminer", "error": str(e), "success": False})
        
        # Method 3: Fallback to PyPDF2
        try:
            page_texts, char_count = self._extract_with_pypdf2(file_path)
            methods_used.append({"method": "pypdf2", "char_count": char_count, "success": True})
            return page_texts, methods_used
        except Exception as e:
            methods_used.append({"method": "pypdf2", "error": str(e), "success": False})
            raise Exception("All text extraction methods failed")
    
    def _extract_with_pymupdf(self, file_path: str, max_pages: Optional[int] = None) -> tuple[List[str], int]:
        """Extract text using PyMuPDF - extracts ALL pages"""
        doc = fitz.open(file_path)
        total_pages = len(doc)
        
        logger.info(f"PyMuPDF: Extracting all {total_pages} pages...")
        page_texts = []
        total_chars = 0
        
        for page_num in range(total_pages):
            page = doc[page_num]
            txt = page.get_text("text")
            page_texts.append(txt)
            total_chars += len(txt)
            
            # Log progress for large documents
            if (page_num + 1) % 25 == 0:
                logger.info(f"PyMuPDF: Extracted {page_num + 1}/{total_pages} pages ({total_chars} chars so far)")
        
        doc.close()
        logger.info(f"PyMuPDF extraction complete: {total_pages} pages, {total_chars} characters")
        return page_texts, total_chars
    
    def _extract_with_pdfminer(self, file_path: str, max_pages: Optional[int] = None) -> tuple[List[str], int]:
        """Extract text using pdfminer - extracts ALL pages"""
        logger.info("pdfminer: Extracting all pages...")
        text = pdfminer_extract(file_path)
        # Split by pages (rough approximation)
        page_texts = text.split('\f')  # Form feed character
        char_count = len(text)
        logger.info(f"pdfminer extraction complete: {len(page_texts)} pages, {char_count} characters")
        return page_texts, char_count
    
    def _extract_with_pypdf2(self, file_path: str, max_pages: Optional[int] = None) -> tuple[List[str], int]:
        """Extract text using PyPDF2 (fallback) - extracts ALL pages"""
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            total_pages = len(pdf_reader.pages)
            
            logger.info(f"PyPDF2: Extracting all {total_pages} pages...")
            page_texts = []
            total_chars = 0
            
            for page_num in range(total_pages):
                page = pdf_reader.pages[page_num]
                txt = page.extract_text()
                page_texts.append(txt)
                total_chars += len(txt)
                
                # Log progress for large documents
                if (page_num + 1) % 25 == 0:
                    logger.info(f"PyPDF2: Extracted {page_num + 1}/{total_pages} pages ({total_chars} chars so far)")
        
        logger.info(f"PyPDF2 extraction complete: {total_pages} pages, {total_chars} characters")
        return page_texts, total_chars
    
    async def _extract_with_ocr(self, file_path: str, trace_dir: Path) -> List[str]:
        """Extract text using OCR (Tesseract)"""
        try:
            # Check if tesseract is available
            subprocess.run(['tesseract', '--version'], capture_output=True, check=True)
            
            # Convert PDF to images
            images_dir = trace_dir / "ocr_images"
            images_dir.mkdir(exist_ok=True)
            
            # Use pdftoppm to convert PDF to images
            subprocess.run([
                'pdftoppm', '-png', '-r', '300', file_path, str(images_dir / "page")
            ], check=True)
            
            # OCR each image
            page_texts = []
            image_files = sorted(images_dir.glob("page-*.png"))
            
            for img_file in image_files:
                result = subprocess.run([
                    'tesseract', str(img_file), 'stdout', '-l', 'eng'
                ], capture_output=True, text=True, check=True)
                page_texts.append(result.stdout)
            
            return page_texts
            
        except (subprocess.CalledProcessError, FileNotFoundError) as e:
            logger.warning(f"OCR failed: {e}")
            return []
    
    def _clean_text_robust(self, page_texts: List[str]) -> str:
        """Robust text cleaning that preserves meaning and structure"""
        # Join all pages
        text = "\n".join(page_texts)
        
        # Fix hyphenation across line breaks: "prohibi-\nted" → "prohibited"
        text = re.sub(r"(\w)-\n(\w)", r"\1\2", text)
        
        # Normalize unicode (fi/ff ligatures, etc.)
        text = unicodedata.normalize("NFKC", text)
        
        # Fix common OCR errors (but preserve newlines)
        # Fix word spacing within lines (not across newlines)
        lines = text.split('\n')
        cleaned_lines = []
        for line in lines:
            # Fix excessive spaces within the line
            line = re.sub(r' +', ' ', line)
            # Fix missing spaces before capitals within the line
            line = re.sub(r"(\w)([A-Z])", r"\1 \2", line)
            cleaned_lines.append(line)
        text = '\n'.join(cleaned_lines)
        
        # Collapse multiple blank lines (but preserve single newlines)
        text = re.sub(r"\n{3,}", "\n\n", text)
        
        # Remove excessive spaces within lines (but preserve newlines)
        # This preserves line structure which is important for table detection
        text = re.sub(r'[ \t]+', ' ', text)  # Collapse spaces/tabs but keep newlines
        
        # Remove page numbers and headers/footers (basic patterns)
        text = re.sub(r'^\s*\d+\s*$', '', text, flags=re.MULTILINE)
        text = re.sub(r'Page \d+ of \d+', '', text, flags=re.IGNORECASE)
        
        return text.strip()
    
    def _check_camelot_dependencies(self) -> Dict[str, bool]:
        """Check if Camelot system dependencies are available"""
        deps = {
            "ghostscript": False,
            "tkinter": False
        }
        
        # Check for Ghostscript
        try:
            result = subprocess.run(
                ["gs", "--version"],
                capture_output=True,
                text=True,
                timeout=5,
                creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0
            )
            if result.returncode == 0:
                deps["ghostscript"] = True
        except (FileNotFoundError, subprocess.TimeoutExpired, Exception):
            pass
        
        # Check for Tkinter (usually available with Python)
        try:
            import tkinter
            deps["tkinter"] = True
        except ImportError:
            pass
        
        return deps
    
    async def _extract_tables(self, file_path: str, trace_dir: Path) -> List[Dict]:
        """Extract tables from PDF using Camelot with both lattice and stream flavors"""
        if not CAMELOT_AVAILABLE:
            logger.debug("Camelot not available - skipping table extraction")
            return []
        
        # Check dependencies once
        deps = self._check_camelot_dependencies()
        if not deps["ghostscript"]:
            logger.debug("Ghostscript not found - Camelot table extraction may fail. Install Ghostscript for table extraction support.")
        
        table_data = []
        table_id = 1
        
        # Method 1: Try lattice first (line-drawn tables)
        try:
            logger.debug("Attempting table extraction with lattice flavor...")
            lattice_tables = camelot.read_pdf(file_path, pages='all', flavor='lattice')
            
            for table in lattice_tables:
                # Convert to markdown-like format for better LLM processing
                markdown_text = self._convert_table_to_markdown(table.df, table_id, table.page, "lattice")
                
                table_dict = {
                    "table_id": table_id,
                    "page": table.page,
                    "accuracy": table.accuracy,
                    "method": "lattice",
                    "data": table.df.to_dict('records'),
                    "text": table.df.to_string(index=False),
                    "markdown": markdown_text
                }
                table_data.append(table_dict)
                table_id += 1
                
            if len(lattice_tables) > 0:
                logger.info(f"Lattice extraction found {len(lattice_tables)} tables")
            
        except Exception as e:
            error_msg = str(e).lower()
            # Provide more specific error messages
            if "file format not supported" in error_msg or "not supported" in error_msg:
                logger.debug(f"Lattice extraction: PDF format not compatible with table detection (this is normal for some PDFs)")
            elif "ghostscript" in error_msg or "gs" in error_msg:
                logger.warning(f"Lattice extraction failed: Ghostscript not found. Install Ghostscript for table extraction.")
            else:
                logger.debug(f"Lattice table extraction failed: {type(e).__name__} - {str(e)[:100]}")
        
        # Method 2: Try stream (whitespace-separated tables)
        try:
            logger.debug("Attempting table extraction with stream flavor...")
            stream_tables = camelot.read_pdf(file_path, pages='all', flavor='stream')
            
            for table in stream_tables:
                # Convert to markdown-like format for better LLM processing
                markdown_text = self._convert_table_to_markdown(table.df, table_id, table.page, "stream")
                
                table_dict = {
                    "table_id": table_id,
                    "page": table.page,
                    "accuracy": table.accuracy,
                    "method": "stream",
                    "data": table.df.to_dict('records'),
                    "text": table.df.to_string(index=False),
                    "markdown": markdown_text
                }
                table_data.append(table_dict)
                table_id += 1
                
            if len(stream_tables) > 0:
                logger.info(f"Stream extraction found {len(stream_tables)} tables")
            
        except Exception as e:
            error_msg = str(e).lower()
            # Provide more specific error messages
            if "file format not supported" in error_msg or "not supported" in error_msg:
                logger.debug(f"Stream extraction: PDF format not compatible with table detection (this is normal for some PDFs)")
            elif "ghostscript" in error_msg or "gs" in error_msg:
                logger.warning(f"Stream extraction failed: Ghostscript not found. Install Ghostscript for table extraction.")
            else:
                logger.debug(f"Stream table extraction failed: {type(e).__name__} - {str(e)[:100]}")
        
        if len(table_data) == 0:
            logger.debug("No tables extracted from PDF (this is normal if PDF has no extractable tables or is image-based)")
        else:
            logger.info(f"Total tables extracted: {len(table_data)}")
        
        return table_data
    
    def _convert_table_to_markdown(self, df, table_id: int, page: int, method: str) -> str:
        """Convert table DataFrame to markdown-like format for better LLM processing"""
        try:
            # Create markdown table with proper formatting
            markdown_lines = []
            markdown_lines.append(f"=== TABLE {table_id} (page {page}) - {method.upper()} ===")
            
            # Convert DataFrame to markdown format
            if not df.empty:
                # Get column headers
                headers = df.columns.tolist()
                markdown_lines.append(" | ".join(str(col) for col in headers))
                markdown_lines.append(" | ".join("---" for _ in headers))
                
                # Add data rows
                for _, row in df.iterrows():
                    row_data = [str(cell) if pd.notna(cell) else "" for cell in row]
                    markdown_lines.append(" | ".join(row_data))
            
            markdown_lines.append(f"=== END TABLE {table_id} ===")
            return "\n".join(markdown_lines)
            
        except Exception as e:
            logger.error(f"Error converting table {table_id} to markdown: {e}", exc_info=True)
            # Fallback to simple text format
            return f"=== TABLE {table_id} (page {page}) - {method.upper()} ===\n{df.to_string(index=False)}\n=== END TABLE {table_id} ==="
    
    def _stitch_tables_into_text(self, text: str, tables: List[Dict]) -> str:
        """Stitch extracted tables back into text with clear markers"""
        if not tables:
            return text
        
        table_texts = []
        for table in tables:
            # Use the markdown format if available, otherwise fall back to text
            if 'markdown' in table:
                table_texts.append(f"\n\n{table['markdown']}\n")
            else:
                # Fallback to old format
                table_marker = f"\n\n[TABLE {table['table_id']} - Page {table['page']}]\n"
                table_content = table['text']
                table_marker += table_content
                table_marker += f"\n[END TABLE {table['table_id']}]\n\n"
                table_texts.append(table_marker)
        
        # Insert tables at the end of the document
        return text + "\n\n" + "\n".join(table_texts)
    
    def _clean_text(self, text: str) -> str:
        """Clean and normalize text"""
        import re
        
        # Remove excessive whitespace
        text = re.sub(r'\s+', ' ', text)
        
        # Remove page numbers and headers/footers (basic patterns)
        text = re.sub(r'^\s*\d+\s*$', '', text, flags=re.MULTILINE)
        text = re.sub(r'Page \d+ of \d+', '', text, flags=re.IGNORECASE)
        
        # Normalize line breaks
        text = re.sub(r'\n\s*\n', '\n\n', text)
        
        return text.strip()
    
    def chunk_text(self, text: str) -> List[Dict[str, Any]]:
        """Create text chunks using LangChain's RecursiveCharacterTextSplitter for better granularity"""
        if not LANGCHAIN_AVAILABLE:
            # Fallback to existing chunking method
            return self._create_chunks(text, max_tokens=1000, overlap_tokens=150)
        
        splitter = RecursiveCharacterTextSplitter(
            chunk_size=1000,        # ~700–1 000 tokens per chunk
            chunk_overlap=150,      # 15% overlap keeps context
            separators=["\n\n", "\n", ". ", ";", " "]
        )
        chunks = splitter.split_text(text)
        
        # Convert to the expected format
        result = []
        char_position = 0
        
        for i, chunk_text in enumerate(chunks):
            chunk_length = len(chunk_text)
            result.append({
                "chunk_id": i + 1,
                "start_char": char_position,
                "end_char": char_position + chunk_length,
                "text": chunk_text,
                "length": chunk_length,
                "token_estimate": chunk_length // 4,
                "prev_chunk": i if i > 0 else None,
                "next_chunk": i + 2 if i < len(chunks) - 1 else None,
                "has_negations": bool(self.NEG_CUES.search(chunk_text)),
                "type": "text"
            })
            char_position += chunk_length
        
        return result
    
    def _create_chunks(self, text: str, max_tokens: int = 1200, overlap_tokens: int = 150) -> List[Dict[str, Any]]:
        """Create text chunks with negation-aware chunking that preserves exceptions and legal constructs"""
        
        # Fallback to simple chunking if NLTK is not available
        if not NLTK_AVAILABLE:
            return self._create_chunks_fallback(text, max_tokens * 4, overlap_tokens * 4)
        
        # Convert token estimates to character estimates (rough approximation: 1 token ≈ 4 characters)
        max_chars = max_tokens * 4
        overlap_chars = overlap_tokens * 4
        
        # Split text into sentences
        sentences = sent_tokenize(text)
        chunks = []
        current_chunk = []
        current_length = 0
        chunk_id = 0
        char_position = 0
        
        i = 0
        while i < len(sentences):
            sentence = sentences[i].strip()
            if not sentence:
                i += 1
                continue
                
            # Check if sentence contains negation cues
            has_negation = bool(self.NEG_CUES.search(sentence))
            
            # Calculate length of current sentence plus potential neighbors
            sentence_length = len(sentence)
            extended_length = sentence_length
            
            # If sentence has negation cues, try to include neighbors
            if has_negation:
                # Add previous sentence if available
                if i > 0 and sentences[i-1].strip():
                    prev_sentence = sentences[i-1].strip()
                    extended_length += len(prev_sentence) + 1  # +1 for space
                
                # Add next sentence if available
                if i + 1 < len(sentences) and sentences[i+1].strip():
                    next_sentence = sentences[i+1].strip()
                    extended_length += len(next_sentence) + 1  # +1 for space
            
            # Check if adding this sentence (with potential neighbors) would exceed limit
            if current_length + extended_length > max_chars and current_chunk:
                # Save current chunk
                chunk_id += 1
                chunk_text = " ".join(current_chunk).strip()
                chunks.append({
                    "chunk_id": chunk_id,
                    "start_char": char_position - current_length,
                    "end_char": char_position,
                    "text": chunk_text,
                    "length": current_length,
                    "token_estimate": current_length // 4,
                    "prev_chunk": chunk_id - 1 if chunk_id > 1 else None,
                    "next_chunk": None,  # Will be updated
                    "has_negations": any(self.NEG_CUES.search(s) for s in current_chunk)
                })
                
                # Create overlap for next chunk
                overlap_text = chunk_text[-overlap_chars:] if len(chunk_text) > overlap_chars else chunk_text
                current_chunk = [overlap_text] if overlap_text else []
                current_length = len(overlap_text)
            else:
                # Add sentence to current chunk
                if has_negation and i > 0 and sentences[i-1].strip() and not current_chunk:
                    # Include previous sentence for context
                    current_chunk.append(sentences[i-1].strip())
                    current_length += len(sentences[i-1].strip()) + 1
                
                current_chunk.append(sentence)
                current_length += sentence_length + (1 if current_chunk else 0)
                
                # If we included next sentence for negation context, skip it in next iteration
                if has_negation and i + 1 < len(sentences) and sentences[i+1].strip():
                    current_chunk.append(sentences[i+1].strip())
                    current_length += len(sentences[i+1].strip()) + 1
                    i += 1  # Skip next sentence as we already included it
            
            char_position += sentence_length + 1  # +1 for space
            i += 1
        
        # Add final chunk
        if current_chunk:
            chunk_id += 1
            chunk_text = " ".join(current_chunk).strip()
            chunks.append({
                "chunk_id": chunk_id,
                "start_char": char_position - current_length,
                "end_char": char_position,
                "text": chunk_text,
                "length": current_length,
                "token_estimate": current_length // 4,
                "prev_chunk": chunk_id - 1 if chunk_id > 1 else None,
                "next_chunk": None,
                "has_negations": any(self.NEG_CUES.search(s) for s in current_chunk)
            })
        
        # Update next_chunk references
        for i in range(len(chunks) - 1):
            chunks[i]["next_chunk"] = chunks[i + 1]["chunk_id"]
        
        return chunks
    
    def _create_chunks_fallback(self, text: str, chunk_size: int = 2000, overlap: int = 200) -> List[Dict[str, Any]]:
        """Fallback chunking method when NLTK is not available"""
        # Split by paragraphs first
        paragraphs = re.split(r'\n\s*\n', text)
        
        chunks = []
        current_chunk = ""
        current_start = 0
        chunk_id = 0
        
        for para in paragraphs:
            if len(current_chunk) + len(para) > chunk_size and current_chunk:
                # Save current chunk
                chunk_id += 1
                chunks.append({
                    "chunk_id": chunk_id,
                    "start_char": current_start,
                    "end_char": current_start + len(current_chunk),
                    "text": current_chunk.strip(),
                    "length": len(current_chunk),
                    "token_estimate": len(current_chunk) // 4,
                    "prev_chunk": chunk_id - 1 if chunk_id > 1 else None,
                    "next_chunk": None,  # Will be updated
                    "has_negations": bool(self.NEG_CUES.search(current_chunk))
                })
                
                # Start new chunk with overlap
                overlap_text = current_chunk[-overlap:] if len(current_chunk) > overlap else current_chunk
                current_chunk = overlap_text + " " + para
                current_start = current_start + len(current_chunk) - len(overlap_text) - len(para) - 1
            else:
                current_chunk += "\n\n" + para if current_chunk else para
        
        # Add final chunk
        if current_chunk.strip():
            chunk_id += 1
            chunks.append({
                "chunk_id": chunk_id,
                "start_char": current_start,
                "end_char": current_start + len(current_chunk),
                "text": current_chunk.strip(),
                "length": len(current_chunk),
                "token_estimate": len(current_chunk) // 4,
                "prev_chunk": chunk_id - 1 if chunk_id > 1 else None,
                "next_chunk": None,
                "has_negations": bool(self.NEG_CUES.search(current_chunk))
            })
        
        # Update next_chunk references
        for i in range(len(chunks) - 1):
            chunks[i]["next_chunk"] = chunks[i + 1]["chunk_id"]
        
        return chunks
    
    async def save_chunks_jsonl(self, chunks: List[Dict[str, Any]], trace_id: str) -> str:
        """Save chunks to JSONL format for verification and analysis"""
        try:
            jsonl_path = os.path.join(self.trace_handler.get_trace_dir(trace_id), "chunks.jsonl")
            
            with open(jsonl_path, 'w', encoding='utf-8') as f:
                for chunk in chunks:
                    # Create a clean JSONL entry
                    jsonl_entry = {
                        "chunk_id": chunk["chunk_id"],
                        "start_char": chunk["start_char"],
                        "end_char": chunk["end_char"],
                        "length": chunk["length"],
                        "token_estimate": chunk.get("token_estimate", chunk["length"] // 4),
                        "has_negations": chunk.get("has_negations", False),
                        "prev_chunk": chunk.get("prev_chunk"),
                        "next_chunk": chunk.get("next_chunk"),
                        "text": chunk["text"]
                    }
                    f.write(json.dumps(jsonl_entry, ensure_ascii=False) + '\n')
            
            return jsonl_path
            
        except Exception as e:
            logger.error(f"Failed to save chunks JSONL: {e}", exc_info=True)
            return None
    
    async def create_excel_export(self, data: dict) -> str:
        """Create Excel export from analysis results"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "OCRD Results"
            
            # Define styles
            header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
            white_font = Font(color="FFFFFF", bold=True)
            center_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Add headers
            headers = ["Section", "Instrument", "Allowed", "Note", "Evidence"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = white_font
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # Add data
            row = 2
            for section, items in data.get("sections", {}).items():
                for key, value in items.items():
                    if isinstance(value, dict) and "allowed" in value:
                        ws.cell(row=row, column=1, value=section)
                        ws.cell(row=row, column=2, value=key)
                        ws.cell(row=row, column=3, value="✓" if value.get("allowed") else "")
                        ws.cell(row=row, column=4, value=value.get("note", ""))
                        # Only show evidence for allowed items
                        evidence_text = ""
                        if value.get("allowed"):
                            evidence_text = value.get("evidence", {}).get("text", "")
                        ws.cell(row=row, column=5, value=evidence_text)
                        
                        # Add borders
                        for col in range(1, 6):
                            ws.cell(row=row, column=col).border = thin_border
                        
                        row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            filename = f"ocrd_export_{int(time.time())}.xlsx"
            file_path = os.path.join(self.export_dir, filename)
            wb.save(file_path)
            
            return file_path
            
        except Exception as e:
            raise Exception(f"Failed to create Excel export: {str(e)}")
    
    def convert_text_to_markdown(self, text: str, filename: str = None) -> str:
        """
        Convert plain text to markdown format with improved structure preservation.
        Preserves markdown table structure when detected.
        
        Args:
            text: The plain text to convert
            filename: Optional filename for the markdown file (without extension)
            
        Returns:
            Markdown formatted string
        """
        if not text:
            return ""
        
        # Create markdown content with proper formatting
        markdown_content = f"# Document Content\n\n"
        
        # Split text into lines for better processing
        lines = text.split('\n')
        current_paragraph = []
        in_table = False
        in_marked_table = False  # Table with explicit markers (=== TABLE ===)
        current_table = []
        
        for i, line in enumerate(lines):
            line = line.rstrip()  # Remove trailing whitespace but preserve leading
            original_line = line  # Keep original for table preservation
            
            # Check if this is a table marker (start or end)
            is_table_marker = self._is_table_marker(line)
            
            # Check if this is a table line (markdown table syntax)
            is_table_line = self._is_table_line(line)
            
            # If we hit a table marker, handle table block
            if is_table_marker:
                # Flush any pending paragraph first
                if current_paragraph:
                    paragraph_text = '\n'.join(current_paragraph).strip()
                    if paragraph_text:
                        if self._is_heading(paragraph_text):
                            markdown_content += f"## {paragraph_text}\n\n"
                        else:
                            markdown_content += f"{paragraph_text}\n\n"
                    current_paragraph = []
                
                # Start or end marked table block
                if not in_marked_table:
                    # Starting a marked table block
                    in_marked_table = True
                    in_table = True
                    current_table = [original_line]
                else:
                    # Ending a marked table block
                    current_table.append(original_line)
                    # Preserve the entire table block as-is
                    markdown_content += '\n'.join(current_table) + '\n\n'
                    current_table = []
                    in_table = False
                    in_marked_table = False
                continue
            
            # If we're inside a marked table block, preserve all lines as-is
            if in_marked_table:
                current_table.append(original_line)
                continue
            
            # If this is a markdown table line (but not a marker), start table block
            if is_table_line and not in_table:
                # Flush any pending paragraph first
                if current_paragraph:
                    paragraph_text = '\n'.join(current_paragraph).strip()
                    if paragraph_text:
                        if self._is_heading(paragraph_text):
                            markdown_content += f"## {paragraph_text}\n\n"
                        else:
                            markdown_content += f"{paragraph_text}\n\n"
                    current_paragraph = []
                
                # Start table block (detected by pipe syntax)
                in_table = True
                current_table = [original_line]
                continue
            
            # If we were in a pipe-detected table but hit a non-table line, end the table
            # Allow one empty line within tables, but end on second empty line or non-table content
            if in_table and not in_marked_table:
                if not line:
                    # Empty line - allow it if we have table content
                    if current_table and len(current_table) > 0:
                        # Check if last line was also empty (two empty lines = end table)
                        if len(current_table) > 1 and not current_table[-1].strip():
                            # Second empty line - end table
                            if current_table:
                                markdown_content += '\n'.join(current_table) + '\n\n'
                                current_table = []
                            in_table = False
                            # Continue to process this empty line as paragraph break
                        else:
                            # First empty line - allow it in table
                            current_table.append(original_line)
                            continue
                    else:
                        # No table content yet - shouldn't happen, but end table
                        in_table = False
                elif not is_table_line:
                    # Non-table, non-empty line - end the table first
                    if current_table:
                        markdown_content += '\n'.join(current_table) + '\n\n'
                        current_table = []
                    in_table = False
                    # Fall through to process this line as regular content
                else:
                    # Still a table line - continue collecting
                    current_table.append(original_line)
                    continue
            
            # Empty line indicates paragraph break
            if not line:
                if current_paragraph:
                    paragraph_text = '\n'.join(current_paragraph).strip()
                    if paragraph_text:
                        # Check if it looks like a heading
                        if self._is_heading(paragraph_text):
                            markdown_content += f"## {paragraph_text}\n\n"
                        else:
                            markdown_content += f"{paragraph_text}\n\n"
                    current_paragraph = []
                continue
            
            # Check if line itself looks like a heading
            if self._is_heading(line) and not current_paragraph:
                # Standalone heading
                markdown_content += f"## {line}\n\n"
            else:
                current_paragraph.append(line)
        
        # Handle any remaining table
        if in_table and current_table:
            markdown_content += '\n'.join(current_table) + '\n\n'
        
        # Add any remaining paragraph
        if current_paragraph:
            paragraph_text = '\n'.join(current_paragraph).strip()
            if paragraph_text:
                if self._is_heading(paragraph_text):
                    markdown_content += f"## {paragraph_text}\n\n"
                else:
                    markdown_content += f"{paragraph_text}\n\n"
        
        return markdown_content
    
    def _is_heading(self, text: str) -> bool:
        """
        Determine if text looks like a heading.
        
        Args:
            text: Text to check
            
        Returns:
            True if text appears to be a heading
        """
        if not text or len(text) > 200:
            return False
        
        # Check for common heading patterns
        # All caps (but not too long)
        if text.isupper() and len(text) < 100:
            return True
        
        # Ends with colon (common in structured documents)
        if text.endswith(':'):
            return True
        
        # Starts with number followed by period or parenthesis (numbered sections)
        if re.match(r'^\d+[\.\)]\s+', text):
            return True
        
        # Starts with letter followed by period (e.g., "A. Section")
        if re.match(r'^[A-Z][\.\)]\s+', text):
            return True
        
        # Common section headers in German documents
        german_headers = [
            'zulässige anlagen', 'unzulässige anlagen', 'anlagegrundsätze',
            'investment policy', 'restrictions', 'permitted', 'prohibited'
        ]
        text_lower = text.lower()
        if any(header in text_lower for header in german_headers):
            return True
        
        return False
    
    def _is_table_line(self, line: str) -> bool:
        """
        Determine if a line is part of a markdown table.
        
        Args:
            line: Line to check
            
        Returns:
            True if line appears to be a markdown table row
        """
        if not line:
            return False
        
        # Check for markdown table syntax: lines with pipe separators
        # Must have at least 2 pipe characters (|) to be a table row
        pipe_count = line.count('|')
        if pipe_count >= 2:
            return True
        
        # Check for table separator row (---|---|---)
        if re.match(r'^[\s\-\|]+$', line.strip()) and '---' in line:
            return True
        
        return False
    
    def _is_table_marker(self, line: str) -> bool:
        """
        Determine if a line is a table marker (start/end of table block).
        
        Args:
            line: Line to check
            
        Returns:
            True if line is a table marker
        """
        if not line:
            return False
        
        line_upper = line.upper().strip()
        
        # Check for table markers like "=== TABLE X ===" or "[TABLE X]"
        if re.match(r'^===?\s*(TABLE|END\s+TABLE)', line_upper):
            return True
        
        if re.match(r'^\[(TABLE|END\s+TABLE)', line_upper):
            return True
        
        return False
    
    async def save_text_as_markdown(self, text: str, job_id: str = None, filename: str = None) -> str:
        """
        Convert text to markdown and save it to the markdown directory.
        
        Args:
            text: The plain text to convert and save
            job_id: Optional job ID to include in filename
            filename: Optional custom filename (without extension)
            
        Returns:
            Path to the saved markdown file
        """
        try:
            # Generate filename if not provided
            if not filename:
                if job_id:
                    filename = f"document_{job_id}"
                else:
                    filename = f"document_{uuid4().hex[:8]}"
            
            # Ensure filename doesn't have extension
            filename = filename.replace('.md', '').replace('.markdown', '')
            
            # Convert to markdown
            markdown_content = self.convert_text_to_markdown(text, filename)
            
            # Create full path
            markdown_path = os.path.join(self.markdown_dir, f"{filename}.md")
            
            # Save markdown file
            async with aiofiles.open(markdown_path, 'w', encoding='utf-8') as f:
                await f.write(markdown_content)
            
            logger.info(f"✅ Markdown file saved: {markdown_path} ({len(markdown_content)} chars)")
            return markdown_path
            
        except Exception as e:
            logger.error(f"❌ Failed to save markdown file: {e}", exc_info=True)
            raise Exception(f"Failed to save markdown file: {str(e)}")
    
    async def read_markdown_file(self, markdown_path: str) -> str:
        """
        Read markdown file content.
        
        Args:
            markdown_path: Path to the markdown file
            
        Returns:
            Content of the markdown file as string
        """
        try:
            async with aiofiles.open(markdown_path, 'r', encoding='utf-8') as f:
                content = await f.read()
            logger.info(f"✅ Markdown file read: {markdown_path} ({len(content)} chars)")
            return content
        except Exception as e:
            logger.error(f"❌ Failed to read markdown file: {e}", exc_info=True)
            raise Exception(f"Failed to read markdown file: {str(e)}")
    
    def cleanup_file(self, file_path: str):
        """Clean up temporary file"""
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
        except:
            pass
